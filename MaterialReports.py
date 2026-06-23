from MaterialContainer import ContinuousMaterialManager
import openpyxl
import os


class MaterialReporter:
    def __init__(self, material_app=None):
        if material_app is None:
            material_app = ContinuousMaterialManager()
            material_app.load_json()
        self.mat_app = material_app

        self.template_path = "Resources/ReportTemplate.xlsx"
        self.save_path = 'temp'

    def generate_new_save_path(self):
        count = 0
        name_list = os.listdir(self.save_path)
        while True:
            name = f"Report {count:010}.xlsx"
            path = f"{self.save_path}/{name}"
            if name in name_list:
                count += 1
            else:
                return path

    def basic_site_report(self, site_id):
        site_obj = self.mat_app.find_site(site_id)
        if site_obj is None:
            raise KeyError(f"Could not find site: {site_id}")

        item_ids = site_obj.list_item_ids()
        report_dict = {item_id: site_obj.count_material(item_id) for item_id in item_ids}
        report_dict = {self.mat_app.lookup(key).item_id: val for key, val in report_dict.items()}

        report_path = self.compile_counts(site_id=site_obj.site_id, count_dict=report_dict)
        print(report_path)

    def compile_counts(self, site_id, count_dict):
        save_path = self.generate_new_save_path()
        book = openpyxl.load_workbook(self.template_path)
        sheet = book.active

        sheet.cell(1, 2).value = site_id

        pointer = 4
        for item_id in sorted(count_dict.keys()):
            count = count_dict[item_id]
            sheet.cell(pointer, 1).value = item_id
            sheet.cell(pointer, 2).value = count
            pointer += 1

        book.save(save_path)
        book.close()

        return save_path

    def clear_temp(self):
        for file_name in os.listdir("temp"):
            os.remove(f"temp/{file_name}")


if __name__ == '__main__':
    reporter = MaterialReporter()
    reporter.clear_temp()
    reporter.basic_site_report('baltimore')
